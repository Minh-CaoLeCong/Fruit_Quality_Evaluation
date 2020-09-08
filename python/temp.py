# export results to excel file
from openpyxl import load_workbook
import xlsxwriter
# using OpenCV library to processing image
import cv2
import numpy as np
# estimate time processing
from timeit import default_timer as timer
import time
# path
import os

########################################################

# RGB
def RGB():
    # the timing of processing
    Start_Time = timer()

    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise two thresholds, a lower (ThresholdCanny1) 
    # and an upper (ThresholdCanny2) of canny edge detector.
    """
    If a pixel has a gradient larger than the upper threshold,
    then it is accepted as an edge pixel; if a pixel is below
    the lower threshold, it is rejected. 
    If the pixel’s gradient is between the thresholds, 
    then it will be accepted only if it is connected to 
    a pixel that is above the high threshold.
    NOTE: Canny recommended a ratio of high:low threshold
            between 2:1 and 3:1
    """
    ThresholdCanny1 = 30
    ThresholdCanny2 = 3 * ThresholdCanny1
    print("Threshold Canny 1: " + str(ThresholdCanny1))
    print("Threshold Canny 2: " + str(ThresholdCanny2))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))

    # load image and convert to grayscale
    OriginalImage_FileName = 'Mango_05_A.JPG'
    print('File Name: ', OriginalImage_FileName)
    OriginalImage_Path = 'E:/CVS/Project/Fruits_Classification/dataset/Mango/' + OriginalImage_FileName
    GrayScale_Image = cv2.imread(OriginalImage_Path, cv2.IMREAD_GRAYSCALE)
    # display image
    cv2.imshow('GrayScale_Image', GrayScale_Image)

    M = GrayScale_Image.shape[0] #Height of image
    N = GrayScale_Image.shape[1] #Width of image

    # smooth image and reduce noise of background
    Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
    # display image
    cv2.imshow('Blur_Image', Blur_Image)

    # using canny algorithm to find out the edge
    Canny_Image = cv2.Canny(Blur_Image, ThresholdCanny1, ThresholdCanny2)
    # display image
    cv2.imshow('Canny_Image', Canny_Image)

    # using dilate operation to try to find connected edge components
    Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
    # display image
    cv2.imshow('Dilate', Dilate_Image)

    # using FloodFill to remove noise
    FloodFill_Image = Dilate_Image
    Gray_FloodFill = Gray_FloodFill_Ini
    Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
    count = 0
    for x in range(0, M):
        for y in range(0, N):
            r = Dilate_Image[x, y]
            if (r == L - 1):
                count += 1
                cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                # FloodFill_Image_Inv = cv2.bitwise_not(FloodFill_Image)
                Gray_FloodFill += Gray_FloodFill_Increase_Step
    # display image
    cv2.imshow('FloodFill_Image', FloodFill_Image)
    FloodFill_Area_Components = [0] * L
    for x in range(0, M):
        for y in range(0, N):
            r = FloodFill_Image[x, y]
            if (r > 0):
                FloodFill_Area_Components[r] += 1
    # print(FloodFill_Area_Components)
    FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
    # print(FloodFill_Area_Components_Enumerate)
    # sorting 'FloodFill_Area_Components_Enumerate' from max to min
    FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
    print("FloodFill Area Components MaxMin: \n" + str(FloodFill_Area_Components_MaxMin))
    rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
    RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
    # remove unnecessary components (noise)
    for x in range(0, M):
        for y in range(0, N):
            r = FloodFill_Image[x, y]
            if (r == rmax):
                RemoveNoise_Image[x, y] = L - 1
            else:
                RemoveNoise_Image[x, y] = 0
    # display image
    cv2.imshow('RemoveNoise_Image', RemoveNoise_Image)

    # find contours
    """
    Because canny edge detector separate different segments,
    so, using 'findContours' to assemble those edge pixels.
    NOTE: The function 'findContours' computes contours from 
            binary images, so it take images created by 'Canny'
    """
    _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    # find fruit contour
    Number_of_Contours = len(Contours)
    # print(Number_of_Contours)
    # using bubble sort to sorting contours from max to min
    Contours_Sorting_MaxMin = Contours.copy()
    for i in range(Number_of_Contours):
        for j in range(0, Number_of_Contours - i - 1):
            if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
    # print("Length Contours:")
    # for i in range(Number_of_Contours):
    #     print("\t\t\t\t\t"+ str(len(Contours[i])))
    print("Length Contours Sorting MaxMin:")
    for i in range(Number_of_Contours):
        print("\t\t\t\t\t"+ str(len(Contours_Sorting_MaxMin[i])))
    # draw fruit contour 0
    DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
    cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
    # display image
    cv2.imshow('DrawFruitContour_Image', DrawFruitContour_Image)

    # estimate fruit area
    Fruit_Area = 0
    Fruit_Area_Outside = 0
    Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
    # throughout the pixel image to check it belongs inside contour or not
    for x in range (0, M):
        for y in range (0, N):
            # using pointPolygonTest to testing if a point is inside contour.
            # the return values are simply +1, -1, or 0 
            # depending on whether the point is inside, out-side, 
            # or on an edge.
            if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                Fruit_Area += 1
                Area_Image[x, y] = L - 1
            else:
                Fruit_Area_Outside += 1
    print("Area: " + str(Fruit_Area))
    print("Area Outside: " + str(Fruit_Area_Outside))
    # display image
    cv2.imshow('Area_Image', Area_Image)

    # Measure fruit dimensions
    # drawing rectangle around contour
    Dimensions_Image = DrawFruitContour_Image
    (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
    cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
    # display image
    cv2.imshow('Dimensions_Image', Dimensions_Image)
    print("Width rectangle: " + str(W_Fruit_Contour))
    print("Height rectangle: " + str(H_Fruit_Contour))

    Stop_Time = timer() - Start_Time
    print("Time: " + str(Stop_Time))

    cv2.waitKey(0)

# Loop, excel
def RGB_Loop_Excel():
    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise two thresholds, a lower (ThresholdCanny1) 
    # and an upper (ThresholdCanny2) of canny edge detector.
    """
    If a pixel has a gradient larger than the upper threshold,
    then it is accepted as an edge pixel; if a pixel is below
    the lower threshold, it is rejected. 
    If the pixel’s gradient is between the thresholds, 
    then it will be accepted only if it is connected to 
    a pixel that is above the high threshold.
    NOTE: Canny recommended a ratio of high:low threshold
            between 2:1 and 3:1
    """
    ThresholdCanny1 = 30
    ThresholdCanny2 = 50
    print("Threshold Canny 1: " + str(ThresholdCanny1))
    print("Threshold Canny 2: " + str(ThresholdCanny2))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))
    cv2.waitKey(100)

    # Input and output image folder path
    Input_Folder_Path = "E:/CVS/Project/Fruits_Classification/dataset/DG_Oranges_Scale/"
    Files_Name_List = []
    # the list of image file extensions
    Image_Extention = ['.JPG', '.jpg', '.JPEG', '.jpeg', '.png', '.PNG']
    # r = root, d = directories, f = files
    for r, d, f in os.walk(Input_Folder_Path):
        for file_name in f:
            if any(Each_Image_Extension in file_name for Each_Image_Extension in Image_Extention):
                Files_Name_List.append(file_name)
    # print(Files_Name_List)
    Output_Folder_Blur_Path = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/blur/"
    Output_Folder_Canny_Path = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/canny/"
    Output_Folder_Dilate_Path = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/dilate/"
    Output_Folder_FloodFill_Path = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/floodfill/"
    Output_Folder_RemoveNoise_Path = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/remove_noise/"
    Output_Folder_Contour_Path = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/contour/"
    Output_Folder_Area_Path = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/area/"
    Output_Folder_Dimension_Path = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/dimension/"

    # Excel file path
    Output_Folder_Excel_File = "E:/CVS/Project/Fruits_Classification/image_result/DG_Oranges_Scale3/"
    
    # create workbook
    WorkBook = xlsxwriter.Workbook(Output_Folder_Excel_File + "DG_Oranges_Scale3_results.xlsx")
    
    # create worksheet for writing parameters
    Parameters_WorkSheet = WorkBook.add_worksheet("Parameters")
    # write parameters to worksheet
    Parameters_WorkSheet.write('A1', "Blur Kernel Size: ") 
    Parameters_WorkSheet.write('A2', "Threshold Canny 1: ") 
    Parameters_WorkSheet.write('A3', "Threshold Canny 2: ") 
    Parameters_WorkSheet.write('A4', "Dilate Kernel Size: ") 
    Parameters_WorkSheet.write('A5', "Gray FloodFill: ") 
    Parameters_WorkSheet.write('A6', "Gray FloodFill Increase Num: ") 
    Parameters_WorkSheet.write('A7', "FloodFill Area Components MaxMin Idx 0: ") 
    Parameters_WorkSheet.write('A8', "FloodFill Area Components MaxMin Idx 1: ") 
    Parameters_WorkSheet.write('A9', "Fruit Contours Sorting MaxMin Idx: ") 
    Parameters_Excel_File_List = []
    Parameters_Excel_File_List.append(str(Blur_KernelSize))
    Parameters_Excel_File_List.append(str(ThresholdCanny1))
    Parameters_Excel_File_List.append(str(ThresholdCanny2))
    Parameters_Excel_File_List.append(str(Dilate_KernelSize))
    Parameters_Excel_File_List.append(str(Gray_FloodFill_Ini))
    Parameters_Excel_File_List.append(str(Gray_FloodFill_Increase_Step))
    Parameters_Excel_File_List.append(str(FloodFill_Area_Components_MaxMin_Idx0))
    Parameters_Excel_File_List.append(str(FloodFill_Area_Components_MaxMin_Idx1))
    Parameters_Excel_File_List.append(str(Fruit_Contours_Sorting_MaxMin_Idx))
    # print(Parameters_Excel_File_List)
    Parameters_Row = 0
    Parameters_Column = 1
    for item in Parameters_Excel_File_List:
        # write operation perform 
        Parameters_WorkSheet.write(Parameters_Row, Parameters_Column, item) 
        # incrementing the value of row by one with each iteratons. 
        Parameters_Row += 1 

    # create worksheet for writing measurements
    Measurements_WorkSheet = WorkBook.add_worksheet("Measurements")
    Measurements_WorkSheet.write('B1', "Time")
    Measurements_WorkSheet.write('C1', "Area")
    Measurements_WorkSheet.write('D1', "Contour_Area")
    Measurements_WorkSheet.write('E1', "Perimeter")
    Measurements_WorkSheet.write('F1', "Width")
    Measurements_WorkSheet.write('G1', "Height")
    Measurements_WorkSheet.write('H1', "FloodFill_Area_Components_MaxMin")
    Measurements_WorkSheet.write('I1', "Contours_Sorting_MaxMin")

    for File_Name_Idx, File_Name in enumerate(Files_Name_List):
        # load image and convert to grayscale
        print('File Name:\t{}\t{}/{}'.format(File_Name, File_Name_Idx + 1, len(Files_Name_List)))
        # the timing of processing
        Start_Time = timer()
        GrayScale_Image = cv2.imread(Input_Folder_Path + File_Name, cv2.IMREAD_GRAYSCALE)

        M = GrayScale_Image.shape[0] #Height of image
        N = GrayScale_Image.shape[1] #Width of image

        # smooth image and reduce noise of background
        Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
        # save image
        cv2.imwrite(Output_Folder_Blur_Path + File_Name, Blur_Image)
        
        # using canny algorithm to find out the edge
        Canny_Image = cv2.Canny(Blur_Image, ThresholdCanny1, ThresholdCanny2)
        # save image
        cv2.imwrite(Output_Folder_Canny_Path + File_Name, Canny_Image)

        # using dilate operation to try to find connected edge components
        Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
        # save image
        cv2.imwrite(Output_Folder_Dilate_Path + File_Name, Dilate_Image)

        # using FloodFill to remove noise
        FloodFill_Image = Dilate_Image
        Gray_FloodFill = Gray_FloodFill_Ini
        Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
        count = 0
        for x in range(0, M):
            for y in range(0, N):
                r = Dilate_Image[x, y]
                if (r == L - 1):
                    count += 1
                    cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                    Gray_FloodFill += Gray_FloodFill_Increase_Step
        # save image
        cv2.imwrite(Output_Folder_FloodFill_Path + File_Name, FloodFill_Image)
        FloodFill_Area_Components = [0] * L
        for x in range(0, M):
            for y in range(0, N):
                r = FloodFill_Image[x, y]
                if (r > 0):
                    FloodFill_Area_Components[r] += 1
        FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
        # sorting 'FloodFill_Area_Components_Enumerate' from max to min
        FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
        rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
        RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        for x in range(0, M):
            for y in range(0, N):
                r = FloodFill_Image[x, y]
                if (r == rmax):
                    RemoveNoise_Image[x, y] = L - 1
                else:
                    RemoveNoise_Image[x, y] = 0
        # save image
        cv2.imwrite(Output_Folder_RemoveNoise_Path + File_Name, RemoveNoise_Image)

        # find contours
        """
        Because canny edge detector separate different segments,
        so, using 'findContours' to assemble those edge pixels.
        NOTE: The function 'findContours' computes contours from 
                binary images, so it take images created by 'Canny'
        """
        _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # find fruit contour
        Number_of_Contours = len(Contours)
        # using bubble sort to sorting contours from max to min
        Contours_Sorting_MaxMin = Contours.copy()
        for i in range(Number_of_Contours):
            for j in range(0, Number_of_Contours - i - 1):
                if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                    Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
        # draw fruit contour '0'
        DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
        # save image
        cv2.imwrite(Output_Folder_Contour_Path + File_Name, DrawFruitContour_Image)

        # estimate fruit area
        Fruit_Area = 0
        Fruit_Perimeter = 0
        Fruit_Area_Contour = 0
        Fruit_Perimeter = cv2.arcLength(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], True)
        Fruit_Area_Contour = cv2.contourArea(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
        Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        # throughout the pixel image to check it belongs inside contour or not
        for x in range (0, M):
            for y in range (0, N):
                # using pointPolygonTest to testing if a point is inside contour.
                # the return values are simply +1, -1, or 0 
                # depending on whether the point is inside, out-side, 
                # or on an edge.
                if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                    Fruit_Area += 1
                    Area_Image[x, y] = L - 1
        # save image
        cv2.imwrite(Output_Folder_Area_Path + File_Name, Area_Image)

        # Measure fruit dimensions
        # drawing rectangle around contour
        Dimensions_Image = DrawFruitContour_Image
        (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
        cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
        # save image
        cv2.imwrite(Output_Folder_Dimension_Path + File_Name, Dimensions_Image)

        Stop_Time = timer() - Start_Time

        # write measurements to worksheet
        Measurements_WorkSheet.write(File_Name_Idx + 1, 0, File_Name)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 1, Stop_Time)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 2, Fruit_Area)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 3, Fruit_Area_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 4, Fruit_Perimeter)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 5, W_Fruit_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 6, H_Fruit_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 7, str(FloodFill_Area_Components_MaxMin))
        Measurements_WorkSheet.write(File_Name_Idx + 1, 8, str(len(Contours_Sorting_MaxMin[0]))) 

        cv2.waitKey(1)
    
    # close excel file
    WorkBook.close()

# Loop, Trackbars
def RGB_Loop_Trackbars():
    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))
    cv2.waitKey(100)

    # Input image folder path
    Input_Folder_Path = "E:/CVS/Project/Fruits_Classification/dataset/Mango/"
    Files_Name_List = []
    # the list of image file extensions
    Image_Extention = ['.JPG', '.jpg', '.JPEG', '.jpeg', '.png', '.PNG']
    # get all the file name of images to list
    # r = root, d = directories, f = files
    for r, d, f in os.walk(Input_Folder_Path):
        for file_name in f:
            if any(Each_Image_Extension in file_name for Each_Image_Extension in Image_Extention):
                Files_Name_List.append(file_name)
    # print(Files_Name_List)

    for File_Name_Idx, File_Name in enumerate(Files_Name_List):
        cv2.waitKey(100)
        print("---------------------------------------------------------------------")
        # load image and convert to grayscale
        print('File Name:\t{}\t{}/{}'.format(File_Name, File_Name_Idx + 1, len(Files_Name_List)))
        Original_Image = cv2.imread(Input_Folder_Path + File_Name)
        # display image
        cv2.imshow('Original_Image', Original_Image)

        # create trackbars
        cv2.namedWindow("TrackBars")
        cv2.resizeWindow("TrackBars", 640, 240)
        # trackbar for threshold canny
        cv2.createTrackbar("ThresholdCanny1", "TrackBars", 0, L - 1, empty)
        cv2.createTrackbar("ThresholdCanny2", "TrackBars", 0, L - 1, empty)

        cv2.waitKey(100)

        # press key to next processing
        print("Press 'Enter' key")
        cv2.waitKey(0)
        print("Enter 'n' key for next image")
        print("Enter 'q' key to closing program")
        print("Press 'Enter' key to continuing processing")
        input_key_in_for_loop = input("Press key 1: ")
        if input_key_in_for_loop == "n":
            continue
        elif input_key_in_for_loop == "q":
            break
        elif input_key_in_for_loop == "":
            pass
        
        check = False

        while True:
            # convert to grayscale
            GrayScale_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2GRAY)
            # display image
            cv2.imshow('GrayScale_Image', GrayScale_Image)
            
            M = GrayScale_Image.shape[0] #Height of image
            N = GrayScale_Image.shape[1] #Width of image

            # smooth image and reduce noise of background
            Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
            # display image
            cv2.imshow('Blur_Image', Blur_Image)
            
            # using canny algorithm to find out the edge
            ThresholdCanny1 = cv2.getTrackbarPos("ThresholdCanny1", "TrackBars")
            ThresholdCanny2 = cv2.getTrackbarPos("ThresholdCanny2", "TrackBars")
            Canny_Image = cv2.Canny(Blur_Image, ThresholdCanny1, ThresholdCanny2)
            # display image
            cv2.imshow('Canny_Image', Canny_Image)

            # using dilate operation to try to find connected edge components
            Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
            # display image
            cv2.imshow('Dilate', Dilate_Image)

            if check == False:    
                print("Enter 'c' key to commanding")
                check = True
            if cv2.waitKey(1) & 0xFF == ord('c'):
                print("Enter 'n' key for next image")
                print("Press 'Enter' key to continuing processing")
                input_key_in_while_true = input("Press key 2: ")
                if input_key_in_while_true == "n":
                    cv2.destroyAllWindows()
                    break
                if input_key_in_while_true == "":
                    print("ThresholdCanny1: " + str(ThresholdCanny1))
                    print("ThresholdCanny2: " + str(ThresholdCanny2))
                    # using FloodFill to remove noise
                    FloodFill_Image = Dilate_Image
                    Gray_FloodFill = Gray_FloodFill_Ini
                    Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
                    count = 0
                    for x in range(0, M):
                        for y in range(0, N):
                            r = Dilate_Image[x, y]
                            if (r == L - 1):
                                count += 1
                                cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                                Gray_FloodFill += Gray_FloodFill_Increase_Step
                    # display image
                    cv2.imshow('FloodFill', FloodFill_Image)
                    FloodFill_Area_Components = [0] * L
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r > 0):
                                FloodFill_Area_Components[r] += 1
                    FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
                    # sorting 'FloodFill_Area_Components_Enumerate' from max to min
                    FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
                    print("FloodFill Area Components MaxMin: \n" + str(FloodFill_Area_Components_MaxMin))
                    rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
                    RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # remove unnecessary components (noise)
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r == rmax):
                                RemoveNoise_Image[x, y] = L - 1
                            else:
                                RemoveNoise_Image[x, y] = 0
                    # display image
                    cv2.imshow('RemoveNoise', RemoveNoise_Image)

                    # find contours
                    """
                    Because canny edge detector separate different segments,
                    so, using 'findContours' to assemble those edge pixels.
                    NOTE: The function 'findContours' computes contours from 
                            binary images, so it take images created by 'Canny'
                    """
                    _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    # find fruit contour
                    Number_of_Contours = len(Contours)
                    # using bubble sort to sorting contours from max to min
                    Contours_Sorting_MaxMin = Contours.copy()
                    for i in range(Number_of_Contours):
                        for j in range(0, Number_of_Contours - i - 1):
                            if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                                Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
                    print("Contours_Sorting_MaxMin: " + str(len(Contours_Sorting_MaxMin[0])))
                    # draw fruit contour '0'
                    DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
                    # display image
                    cv2.imshow('DrawFruitContour', DrawFruitContour_Image)

                    # estimate fruit area
                    Fruit_Area = 0
                    Fruit_Perimeter = 0
                    Fruit_Area_Contour = 0
                    Fruit_Perimeter = cv2.arcLength(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], True)
                    Fruit_Area_Contour = cv2.contourArea(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # throughout the pixel image to check it belongs inside contour or not
                    for x in range (0, M):
                        for y in range (0, N):
                            # using pointPolygonTest to testing if a point is inside contour.
                            # the return values are simply +1, -1, or 0 
                            # depending on whether the point is inside, out-side, 
                            # or on an edge.
                            if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                                Fruit_Area += 1
                                Area_Image[x, y] = L - 1
                    print("Fruit_Area: " + str(Fruit_Area))
                    print("Fruit_Area_Contour: " + str(Fruit_Area_Contour))
                    print("Fruit_Perimeter: " + str(Fruit_Perimeter))
                    # display image
                    cv2.imshow('Area', Area_Image)

                    # Measure fruit dimensions
                    # drawing rectangle around contour
                    Dimensions_Image = DrawFruitContour_Image
                    (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
                    print("W_Fruit_Contour: " + str(W_Fruit_Contour))
                    print("H_Fruit_Contour: " + str(H_Fruit_Contour))
                    # display image
                    cv2.imshow('Dimensions_Image', Dimensions_Image)
                    
                    print("Press 'Enter' key")
                    cv2.waitKey(0)
                    cv2.destroyAllWindows()
                    break
                
            cv2.waitKey(1)
########################################################

# HSV
def HSV():
    # the timing of processing
    Start_Time = timer()

    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise two thresholds, a lower (ThresholdCanny1) 
    # and an upper (ThresholdCanny2) of canny edge detector.
    """
    If a pixel has a gradient larger than the upper threshold,
    then it is accepted as an edge pixel; if a pixel is below
    the lower threshold, it is rejected. 
    If the pixel’s gradient is between the thresholds, 
    then it will be accepted only if it is connected to 
    a pixel that is above the high threshold.
    NOTE: Canny recommended a ratio of high:low threshold
            between 2:1 and 3:1
    """
    ThresholdCanny1 = 30
    ThresholdCanny2 = 3 * ThresholdCanny1
    print("Threshold Canny 1: " + str(ThresholdCanny1))
    print("Threshold Canny 2: " + str(ThresholdCanny2))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))

    # load image
    OriginalImage_FileName = 'Orange-ICC_149.JPG'
    print('File Name: ', OriginalImage_FileName)
    OriginalImage_Path = 'E:/CVS/Project/Fruits_Classification/dataset/DB_Royal/' + OriginalImage_FileName
    Original_Image = cv2.imread(OriginalImage_Path)
    # display image
    cv2.imshow('Original_Image', Original_Image)

    # convert BGR to HSV
    HSV_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2HSV)
    # display image
    cv2.imshow('HSV_Image', HSV_Image)

    # convert to grayscale
    GrayScale_Image = cv2.cvtColor(HSV_Image, cv2.COLOR_BGR2GRAY)
    # display image
    cv2.imshow('GrayScale_Image', GrayScale_Image)

    M = GrayScale_Image.shape[0] #Height of image
    N = GrayScale_Image.shape[1] #Width of image

    # smooth image and reduce noise of background
    Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
    # display image
    cv2.imshow('Blur_Image', Blur_Image)

    # using canny algorithm to find out the edge
    Canny_Image = cv2.Canny(Blur_Image, ThresholdCanny1, ThresholdCanny2)
    # display image
    cv2.imshow('Canny_Image', Canny_Image)

    # using dilate operation to try to find connected edge components
    Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
    # display image
    cv2.imshow('Dilate', Dilate_Image)

    # using FloodFill to remove noise
    FloodFill_Image = Dilate_Image
    Gray_FloodFill = Gray_FloodFill_Ini
    Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
    count = 0
    for x in range(0, M):
        for y in range(0, N):
            r = Dilate_Image[x, y]
            if (r == L - 1):
                count += 1
                cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                # FloodFill_Image_Inv = cv2.bitwise_not(FloodFill_Image)
                Gray_FloodFill += Gray_FloodFill_Increase_Step
    # display image
    cv2.imshow('FloodFill_Image', FloodFill_Image)
    FloodFill_Area_Components = [0] * L
    for x in range(0, M):
        for y in range(0, N):
            r = FloodFill_Image[x, y]
            if (r > 0):
                FloodFill_Area_Components[r] += 1
    # print(FloodFill_Area_Components)
    FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
    # print(FloodFill_Area_Components_Enumerate)
    # sorting 'FloodFill_Area_Components_Enumerate' from max to min
    FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
    print("FloodFill Area Components MaxMin: \n" + str(FloodFill_Area_Components_MaxMin))
    rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
    RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
    # remove unnecessary components (noise)
    for x in range(0, M):
        for y in range(0, N):
            r = FloodFill_Image[x, y]
            if (r == rmax):
                RemoveNoise_Image[x, y] = L - 1
            else:
                RemoveNoise_Image[x, y] = 0
    # display image
    cv2.imshow('RemoveNoise_Image', RemoveNoise_Image)

    # find contours
    """
    Because canny edge detector separate different segments,
    so, using 'findContours' to assemble those edge pixels.
    NOTE: The function 'findContours' computes contours from 
            binary images, so it take images created by 'Canny'
    """
    _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    # find fruit contour
    Number_of_Contours = len(Contours)
    # print(Number_of_Contours)
    # using bubble sort to sorting contours from max to min
    Contours_Sorting_MaxMin = Contours.copy()
    for i in range(Number_of_Contours):
        for j in range(0, Number_of_Contours - i - 1):
            if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
    # print("Length Contours:")
    # for i in range(Number_of_Contours):
    #     print("\t\t\t\t\t"+ str(len(Contours[i])))
    print("Length Contours Sorting MaxMin:")
    for i in range(Number_of_Contours):
        print("\t\t\t\t\t"+ str(len(Contours_Sorting_MaxMin[i])))
    # draw fruit contour 0
    DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
    cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
    # display image
    cv2.imshow('DrawFruitContour_Image', DrawFruitContour_Image)

    # estimate fruit area
    Fruit_Area = 0
    Fruit_Area_Outside = 0
    Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
    # throughout the pixel image to check it belongs inside contour or not
    for x in range (0, M):
        for y in range (0, N):
            # using pointPolygonTest to testing if a point is inside contour.
            # the return values are simply +1, -1, or 0 
            # depending on whether the point is inside, out-side, 
            # or on an edge.
            if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                Fruit_Area += 1
                Area_Image[x, y] = L - 1
            else:
                Fruit_Area_Outside += 1
    print("Area: " + str(Fruit_Area))
    print("Area Outside: " + str(Fruit_Area_Outside))
    # display image
    cv2.imshow('Area_Image', Area_Image)

    # Measure fruit dimensions
    # drawing rectangle around contour
    Dimensions_Image = DrawFruitContour_Image
    (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
    cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
    # display image
    cv2.imshow('Dimensions_Image', Dimensions_Image)
    print("Width rectangle: " + str(W_Fruit_Contour))
    print("Height rectangle: " + str(H_Fruit_Contour))

    Stop_Time = timer() - Start_Time
    print("Time: " + str(Stop_Time))

    cv2.waitKey(0)

# Loop, HSV, excel
def HSV_Loop_Excel():
    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise two thresholds, a lower (ThresholdCanny1) 
    # and an upper (ThresholdCanny2) of canny edge detector.
    """
    If a pixel has a gradient larger than the upper threshold,
    then it is accepted as an edge pixel; if a pixel is below
    the lower threshold, it is rejected. 
    If the pixel’s gradient is between the thresholds, 
    then it will be accepted only if it is connected to 
    a pixel that is above the high threshold.
    NOTE: Canny recommended a ratio of high:low threshold
            between 2:1 and 3:1
    """
    ThresholdCanny1 = 50
    ThresholdCanny2 = 2 * ThresholdCanny1
    print("Threshold Canny 1: " + str(ThresholdCanny1))
    print("Threshold Canny 2: " + str(ThresholdCanny2))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))
    cv2.waitKey(100)

    # Input and output image folder path
    Input_Folder_Path = "E:/CVS/Project/Fruits_Classification/dataset/Mango/"
    Files_Name_List = []
    # the list of image file extensions
    Image_Extention = ['.JPG', '.jpg', '.JPEG', '.jpeg', '.png', '.PNG']
    # r = root, d = directories, f = files
    for r, d, f in os.walk(Input_Folder_Path):
        for file_name in f:
            if any(Each_Image_Extension in file_name for Each_Image_Extension in Image_Extention):
                Files_Name_List.append(file_name)
    # print(Files_Name_List)
    Output_Folder_Hsv_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/hsv/"
    Output_Folder_Gray_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/gray/"
    Output_Folder_Blur_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/blur/"
    Output_Folder_Canny_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/canny/"
    Output_Folder_Dilate_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/dilate/"
    Output_Folder_FloodFill_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/floodfill/"
    Output_Folder_RemoveNoise_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/remove_noise/"
    Output_Folder_Contour_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/contour/"
    Output_Folder_Area_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/area/"
    Output_Folder_Dimension_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/dimension/"

    # Excel file path
    Output_Folder_Excel_File = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Mango6/"
    
    # create workbook
    WorkBook = xlsxwriter.Workbook(Output_Folder_Excel_File + "HSV_Mango6_results.xlsx")
    
    # create worksheet for writing parameters
    Parameters_WorkSheet = WorkBook.add_worksheet("Parameters")
    # write parameters to worksheet
    Parameters_WorkSheet.write('A1', "Blur Kernel Size: ") 
    Parameters_WorkSheet.write('A2', "Threshold Canny 1: ") 
    Parameters_WorkSheet.write('A3', "Threshold Canny 2: ") 
    Parameters_WorkSheet.write('A4', "Dilate Kernel Size: ") 
    Parameters_WorkSheet.write('A5', "Gray FloodFill: ") 
    Parameters_WorkSheet.write('A6', "Gray FloodFill Increase Num: ") 
    Parameters_WorkSheet.write('A7', "FloodFill Area Components MaxMin Idx 0: ") 
    Parameters_WorkSheet.write('A8', "FloodFill Area Components MaxMin Idx 1: ") 
    Parameters_WorkSheet.write('A9', "Fruit Contours Sorting MaxMin Idx: ") 
    Parameters_Excel_File_List = []
    Parameters_Excel_File_List.append(str(Blur_KernelSize))
    Parameters_Excel_File_List.append(str(ThresholdCanny1))
    Parameters_Excel_File_List.append(str(ThresholdCanny2))
    Parameters_Excel_File_List.append(str(Dilate_KernelSize))
    Parameters_Excel_File_List.append(str(Gray_FloodFill_Ini))
    Parameters_Excel_File_List.append(str(Gray_FloodFill_Increase_Step))
    Parameters_Excel_File_List.append(str(FloodFill_Area_Components_MaxMin_Idx0))
    Parameters_Excel_File_List.append(str(FloodFill_Area_Components_MaxMin_Idx1))
    Parameters_Excel_File_List.append(str(Fruit_Contours_Sorting_MaxMin_Idx))
    # print(Parameters_Excel_File_List)
    Parameters_Row = 0
    Parameters_Column = 1
    for item in Parameters_Excel_File_List:
        # write operation perform 
        Parameters_WorkSheet.write(Parameters_Row, Parameters_Column, item) 
        # incrementing the value of row by one with each iteratons. 
        Parameters_Row += 1 

    # create worksheet for writing measurements
    Measurements_WorkSheet = WorkBook.add_worksheet("Measurements")
    Measurements_WorkSheet.write('B1', "Time")
    Measurements_WorkSheet.write('C1', "Area")
    Measurements_WorkSheet.write('D1', "Contour_Area")
    Measurements_WorkSheet.write('E1', "Perimeter")
    Measurements_WorkSheet.write('F1', "Width")
    Measurements_WorkSheet.write('G1', "Height")
    Measurements_WorkSheet.write('H1', "FloodFill_Area_Components_MaxMin")
    Measurements_WorkSheet.write('I1', "Contours_Sorting_MaxMin")

    for File_Name_Idx, File_Name in enumerate(Files_Name_List):
        # load image and convert to grayscale
        print('File Name:\t{}\t{}/{}'.format(File_Name, File_Name_Idx + 1, len(Files_Name_List)))
        # the timing of processing
        Start_Time = timer()
        Original_Image = cv2.imread(Input_Folder_Path + File_Name)

        # convert BGR to HSV
        HSV_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2HSV)
        # save image
        cv2.imwrite(Output_Folder_Hsv_Path + File_Name, HSV_Image)

        # convert to grayscale
        GrayScale_Image = cv2.cvtColor(HSV_Image, cv2.COLOR_BGR2GRAY)
        # save image
        cv2.imwrite(Output_Folder_Gray_Path + File_Name, GrayScale_Image)

        M = GrayScale_Image.shape[0] #Height of image
        N = GrayScale_Image.shape[1] #Width of image

        # smooth image and reduce noise of background
        Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
        # save image
        cv2.imwrite(Output_Folder_Blur_Path + File_Name, Blur_Image)
        
        # using canny algorithm to find out the edge
        Canny_Image = cv2.Canny(Blur_Image, ThresholdCanny1, ThresholdCanny2)
        # save image
        cv2.imwrite(Output_Folder_Canny_Path + File_Name, Canny_Image)

        # using dilate operation to try to find connected edge components
        Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
        # save image
        cv2.imwrite(Output_Folder_Dilate_Path + File_Name, Dilate_Image)

        # using FloodFill to remove noise
        FloodFill_Image = Dilate_Image
        Gray_FloodFill = Gray_FloodFill_Ini
        Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
        count = 0
        for x in range(0, M):
            for y in range(0, N):
                r = Dilate_Image[x, y]
                if (r == L - 1):
                    count += 1
                    cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                    Gray_FloodFill += Gray_FloodFill_Increase_Step
        # save image
        cv2.imwrite(Output_Folder_FloodFill_Path + File_Name, FloodFill_Image)
        FloodFill_Area_Components = [0] * L
        for x in range(0, M):
            for y in range(0, N):
                r = FloodFill_Image[x, y]
                if (r > 0):
                    FloodFill_Area_Components[r] += 1
        FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
        # sorting 'FloodFill_Area_Components_Enumerate' from max to min
        FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
        rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
        RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        for x in range(0, M):
            for y in range(0, N):
                r = FloodFill_Image[x, y]
                if (r == rmax):
                    RemoveNoise_Image[x, y] = L - 1
                else:
                    RemoveNoise_Image[x, y] = 0
        # save image
        cv2.imwrite(Output_Folder_RemoveNoise_Path + File_Name, RemoveNoise_Image)

        # find contours
        """
        Because canny edge detector separate different segments,
        so, using 'findContours' to assemble those edge pixels.
        NOTE: The function 'findContours' computes contours from 
                binary images, so it take images created by 'Canny'
        """
        _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # find fruit contour
        Number_of_Contours = len(Contours)
        # using bubble sort to sorting contours from max to min
        Contours_Sorting_MaxMin = Contours.copy()
        for i in range(Number_of_Contours):
            for j in range(0, Number_of_Contours - i - 1):
                if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                    Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
        # draw fruit contour '0'
        DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
        # save image
        cv2.imwrite(Output_Folder_Contour_Path + File_Name, DrawFruitContour_Image)

        # estimate fruit area
        Fruit_Area = 0
        Fruit_Perimeter = 0
        Fruit_Area_Contour = 0
        Fruit_Perimeter = cv2.arcLength(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], True)
        Fruit_Area_Contour = cv2.contourArea(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
        Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        # throughout the pixel image to check it belongs inside contour or not
        for x in range (0, M):
            for y in range (0, N):
                # using pointPolygonTest to testing if a point is inside contour.
                # the return values are simply +1, -1, or 0 
                # depending on whether the point is inside, out-side, 
                # or on an edge.
                if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                    Fruit_Area += 1
                    Area_Image[x, y] = L - 1
        # save image
        cv2.imwrite(Output_Folder_Area_Path + File_Name, Area_Image)

        # Measure fruit dimensions
        # drawing rectangle around contour
        Dimensions_Image = DrawFruitContour_Image
        (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
        cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
        # save image
        cv2.imwrite(Output_Folder_Dimension_Path + File_Name, Dimensions_Image)

        Stop_Time = timer() - Start_Time

        # write measurements to worksheet
        Measurements_WorkSheet.write(File_Name_Idx + 1, 0, File_Name)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 1, Stop_Time)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 2, Fruit_Area)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 3, Fruit_Area_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 4, Fruit_Perimeter)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 5, W_Fruit_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 6, H_Fruit_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 7, str(FloodFill_Area_Components_MaxMin))
        Measurements_WorkSheet.write(File_Name_Idx + 1, 8, str(len(Contours_Sorting_MaxMin[0]))) 

        cv2.waitKey(1)
    
    # close excel file
    WorkBook.close()

# Loop, HSV, Trackbars
def HSV_Loop_Trackbars():
    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))
    cv2.waitKey(100)

    # Input image folder path
    Input_Folder_Path = "E:/CVS/Project/Fruits_Classification/dataset/Mango/"
    Files_Name_List = []
    # the list of image file extensions
    Image_Extention = ['.JPG', '.jpg', '.JPEG', '.jpeg', '.png', '.PNG']
    # get all the file name of images to list
    # r = root, d = directories, f = files
    for r, d, f in os.walk(Input_Folder_Path):
        for file_name in f:
            if any(Each_Image_Extension in file_name for Each_Image_Extension in Image_Extention):
                Files_Name_List.append(file_name)
    # print(Files_Name_List)

    for File_Name_Idx, File_Name in enumerate(Files_Name_List):
        cv2.waitKey(100)
        print("---------------------------------------------------------------------")
        # load image and convert to grayscale
        print('File Name:\t{}\t{}/{}'.format(File_Name, File_Name_Idx + 1, len(Files_Name_List)))
        Original_Image = cv2.imread(Input_Folder_Path + File_Name)
        # display image
        cv2.imshow('Original_Image', Original_Image)

        # create trackbars
        cv2.namedWindow("TrackBars")
        cv2.resizeWindow("TrackBars", 640, 240)
        # trackbar for threshold canny
        cv2.createTrackbar("ThresholdCanny1", "TrackBars", 0, L - 1, empty)
        cv2.createTrackbar("ThresholdCanny2", "TrackBars", 0, L - 1, empty)

        cv2.waitKey(100)

        # press key to next processing
        print("Press 'Enter' key")
        cv2.waitKey(0)
        print("Enter 'n' key for next image")
        print("Enter 'q' key to closing program")
        print("Press 'Enter' key to continuing processing")
        input_key_in_for_loop = input("Press key 1: ")
        if input_key_in_for_loop == "n":
            continue
        elif input_key_in_for_loop == "q":
            break
        elif input_key_in_for_loop == "":
            pass
        
        check = False

        while True:
            # convert BGR to HSV
            HSV_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2HSV)
            # display image
            cv2.imshow('HSV_Image', HSV_Image)

            # convert to grayscale
            GrayScale_Image = cv2.cvtColor(HSV_Image, cv2.COLOR_BGR2GRAY)
            # display image
            cv2.imshow('GrayScale_Image', GrayScale_Image)
            
            M = GrayScale_Image.shape[0] #Height of image
            N = GrayScale_Image.shape[1] #Width of image

            # smooth image and reduce noise of background
            Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
            # display image
            cv2.imshow('Blur_Image', Blur_Image)
            
            # using canny algorithm to find out the edge
            ThresholdCanny1 = cv2.getTrackbarPos("ThresholdCanny1", "TrackBars")
            ThresholdCanny2 = cv2.getTrackbarPos("ThresholdCanny2", "TrackBars")
            Canny_Image = cv2.Canny(Blur_Image, ThresholdCanny1, ThresholdCanny2)
            # display image
            cv2.imshow('Canny_Image', Canny_Image)

            # using dilate operation to try to find connected edge components
            Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
            # display image
            cv2.imshow('Dilate', Dilate_Image)

            if check == False:    
                print("Enter 'c' key to commanding")
                check = True
            if cv2.waitKey(1) & 0xFF == ord('c'):
                print("Enter 'n' key for next image")
                print("Press 'Enter' key to continuing processing")
                input_key_in_while_true = input("Press key 2: ")
                if input_key_in_while_true == "n":
                    cv2.destroyAllWindows()
                    break
                if input_key_in_while_true == "":
                    print("ThresholdCanny1: " + str(ThresholdCanny1))
                    print("ThresholdCanny2: " + str(ThresholdCanny2))
                    # using FloodFill to remove noise
                    FloodFill_Image = Dilate_Image
                    Gray_FloodFill = Gray_FloodFill_Ini
                    Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
                    count = 0
                    for x in range(0, M):
                        for y in range(0, N):
                            r = Dilate_Image[x, y]
                            if (r == L - 1):
                                count += 1
                                cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                                Gray_FloodFill += Gray_FloodFill_Increase_Step
                    # display image
                    cv2.imshow('FloodFill', FloodFill_Image)
                    FloodFill_Area_Components = [0] * L
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r > 0):
                                FloodFill_Area_Components[r] += 1
                    FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
                    # sorting 'FloodFill_Area_Components_Enumerate' from max to min
                    FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
                    print("FloodFill Area Components MaxMin: \n" + str(FloodFill_Area_Components_MaxMin))
                    rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
                    RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # remove unnecessary components (noise)
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r == rmax):
                                RemoveNoise_Image[x, y] = L - 1
                            else:
                                RemoveNoise_Image[x, y] = 0
                    # display image
                    cv2.imshow('RemoveNoise', RemoveNoise_Image)

                    # find contours
                    """
                    Because canny edge detector separate different segments,
                    so, using 'findContours' to assemble those edge pixels.
                    NOTE: The function 'findContours' computes contours from 
                            binary images, so it take images created by 'Canny'
                    """
                    _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    # find fruit contour
                    Number_of_Contours = len(Contours)
                    # using bubble sort to sorting contours from max to min
                    Contours_Sorting_MaxMin = Contours.copy()
                    for i in range(Number_of_Contours):
                        for j in range(0, Number_of_Contours - i - 1):
                            if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                                Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
                    print("Contours_Sorting_MaxMin: " + str(len(Contours_Sorting_MaxMin[0])))
                    # draw fruit contour '0'
                    DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
                    # display image
                    cv2.imshow('DrawFruitContour', DrawFruitContour_Image)

                    # estimate fruit area
                    Fruit_Area = 0
                    Fruit_Perimeter = 0
                    Fruit_Area_Contour = 0
                    Fruit_Perimeter = cv2.arcLength(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], True)
                    Fruit_Area_Contour = cv2.contourArea(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # throughout the pixel image to check it belongs inside contour or not
                    for x in range (0, M):
                        for y in range (0, N):
                            # using pointPolygonTest to testing if a point is inside contour.
                            # the return values are simply +1, -1, or 0 
                            # depending on whether the point is inside, out-side, 
                            # or on an edge.
                            if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                                Fruit_Area += 1
                                Area_Image[x, y] = L - 1
                    print("Fruit_Area: " + str(Fruit_Area))
                    print("Fruit_Area_Contour: " + str(Fruit_Area_Contour))
                    print("Fruit_Perimeter: " + str(Fruit_Perimeter))
                    # display image
                    cv2.imshow('Area', Area_Image)

                    # Measure fruit dimensions
                    # drawing rectangle around contour
                    Dimensions_Image = DrawFruitContour_Image
                    (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
                    print("W_Fruit_Contour: " + str(W_Fruit_Contour))
                    print("H_Fruit_Contour: " + str(H_Fruit_Contour))
                    # display image
                    cv2.imshow('Dimensions_Image', Dimensions_Image)
                    
                    print("Press 'Enter' key")
                    cv2.waitKey(0)
                    cv2.destroyAllWindows()
                    break
                
            cv2.waitKey(1)
########################################################

# inRange
def inRange():
    # the timing of processing
    Start_Time = timer()

    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # # initialise two thresholds, a lower (ThresholdCanny1) 
    # # and an upper (ThresholdCanny2) of canny edge detector.
    # """
    # If a pixel has a gradient larger than the upper threshold,
    # then it is accepted as an edge pixel; if a pixel is below
    # the lower threshold, it is rejected. 
    # If the pixel’s gradient is between the thresholds, 
    # then it will be accepted only if it is connected to 
    # a pixel that is above the high threshold.
    # NOTE: Canny recommended a ratio of high:low threshold
    #         between 2:1 and 3:1
    # """
    # ThresholdCanny1 = 30
    # ThresholdCanny2 = 3 * ThresholdCanny1
    # print("Threshold Canny 1: " + str(ThresholdCanny1))
    # print("Threshold Canny 2: " + str(ThresholdCanny2))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))

    # load image
    OriginalImage_FileName = 'Mango_23_A.JPG'
    print('File Name: ', OriginalImage_FileName)
    OriginalImage_Path = 'E:/CVS/Project/Fruits_Classification/dataset/Mango/' + OriginalImage_FileName
    Original_Image = cv2.imread(OriginalImage_Path)
    # display image
    cv2.imshow('Original_Image', Original_Image)

    # convert to grayscale
    GrayScale_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2GRAY)
    # display image
    cv2.imshow('GrayScale_Image', GrayScale_Image)

    M = GrayScale_Image.shape[0] #Height of image
    N = GrayScale_Image.shape[1] #Width of image

    # smooth image and reduce noise of background
    Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
    # display image
    cv2.imshow('Blur_Image', Blur_Image)

    # create trackbars
    cv2.namedWindow("TrackBars")
    cv2.resizeWindow("TrackBars", 640, 240)
    # trackbar for threshold canny
    cv2.createTrackbar("ThresholdCanny1", "TrackBars", 0, L - 1, empty)
    cv2.createTrackbar("ThresholdCanny2", "TrackBars", 0, L - 1, empty)
    cv2.createTrackbar("inRange1", "TrackBars", 0, L - 1, empty)
    cv2.createTrackbar("inRange2", "TrackBars", 0, L - 1, empty)

    while True:
        ThresholdCanny1 = cv2.getTrackbarPos("ThresholdCanny1", "TrackBars")
        ThresholdCanny2 = cv2.getTrackbarPos("ThresholdCanny2", "TrackBars")
        inRange1 = cv2.getTrackbarPos("inRange1", "TrackBars")
        inRange2 = cv2.getTrackbarPos("inRange2", "TrackBars")

        # inRange
        inRange_Image = cv2.inRange(Blur_Image, inRange1, inRange2)
        # display image
        cv2.imshow('inRange_Image', inRange_Image)

        # using canny algorithm to find out the edge
        Canny_Image = cv2.Canny(inRange_Image, ThresholdCanny1, ThresholdCanny2)
        # display image
        cv2.imshow('Canny_Image', Canny_Image)

        # using dilate operation to try to find connected edge components
        Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
        # display image
        cv2.imshow('Dilate', Dilate_Image)

        if cv2.waitKey(1) & 0xFF == ord('n'):
            # using FloodFill to remove noise
            FloodFill_Image = Dilate_Image
            Gray_FloodFill = Gray_FloodFill_Ini
            Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
            count = 0
            for x in range(0, M):
                for y in range(0, N):
                    r = Dilate_Image[x, y]
                    if (r == L - 1):
                        count += 1
                        cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                        # FloodFill_Image_Inv = cv2.bitwise_not(FloodFill_Image)
                        Gray_FloodFill += Gray_FloodFill_Increase_Step
            # display image
            cv2.imshow('FloodFill_Image', FloodFill_Image)
            FloodFill_Area_Components = [0] * L
            for x in range(0, M):
                for y in range(0, N):
                    r = FloodFill_Image[x, y]
                    if (r > 0):
                        FloodFill_Area_Components[r] += 1
            # print(FloodFill_Area_Components)
            FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
            # print(FloodFill_Area_Components_Enumerate)
            # sorting 'FloodFill_Area_Components_Enumerate' from max to min
            FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
            print("FloodFill Area Components MaxMin: \n" + str(FloodFill_Area_Components_MaxMin))
            rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
            RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
            # remove unnecessary components (noise)
            for x in range(0, M):
                for y in range(0, N):
                    r = FloodFill_Image[x, y]
                    if (r == rmax):
                        RemoveNoise_Image[x, y] = L - 1
                    else:
                        RemoveNoise_Image[x, y] = 0
            # display image
            cv2.imshow('RemoveNoise_Image', RemoveNoise_Image)

            # find contours
            """
            Because canny edge detector separate different segments,
            so, using 'findContours' to assemble those edge pixels.
            NOTE: The function 'findContours' computes contours from 
                    binary images, so it take images created by 'Canny'
            """
            _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            # find fruit contour
            Number_of_Contours = len(Contours)
            # print(Number_of_Contours)
            # using bubble sort to sorting contours from max to min
            Contours_Sorting_MaxMin = Contours.copy()
            for i in range(Number_of_Contours):
                for j in range(0, Number_of_Contours - i - 1):
                    if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                        Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
            # print("Length Contours:")
            # for i in range(Number_of_Contours):
            #     print("\t\t\t\t\t"+ str(len(Contours[i])))
            print("Length Contours Sorting MaxMin:")
            for i in range(Number_of_Contours):
                print("\t\t\t\t\t"+ str(len(Contours_Sorting_MaxMin[i])))
            # draw fruit contour 0
            DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
            cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
            # display image
            cv2.imshow('DrawFruitContour_Image', DrawFruitContour_Image)

            # estimate fruit area
            Fruit_Area = 0
            Fruit_Area_Outside = 0
            Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
            # throughout the pixel image to check it belongs inside contour or not
            for x in range (0, M):
                for y in range (0, N):
                    # using pointPolygonTest to testing if a point is inside contour.
                    # the return values are simply +1, -1, or 0 
                    # depending on whether the point is inside, out-side, 
                    # or on an edge.
                    if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                        Fruit_Area += 1
                        Area_Image[x, y] = L - 1
                    else:
                        Fruit_Area_Outside += 1
            print("Area: " + str(Fruit_Area))
            print("Area Outside: " + str(Fruit_Area_Outside))
            # display image
            cv2.imshow('Area_Image', Area_Image)

            # Measure fruit dimensions
            # drawing rectangle around contour
            Dimensions_Image = DrawFruitContour_Image
            (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
            cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
            # display image
            cv2.imshow('Dimensions_Image', Dimensions_Image)
            print("Width rectangle: " + str(W_Fruit_Contour))
            print("Height rectangle: " + str(H_Fruit_Contour))

            Stop_Time = timer() - Start_Time
            print("Time: " + str(Stop_Time))

            cv2.waitKey(0)
            
        cv2.waitKey(1)

# Loop, inRange, Trackbars
def inRange_Loop_Trackbars():
    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))
    cv2.waitKey(100)

    # Input image folder path
    Input_Folder_Path = "E:/CVS/Project/Fruits_Classification/image_result/HSV/Citrus4/hsv/"
    Files_Name_List = []
    # the list of image file extensions
    Image_Extention = ['.JPG', '.jpg', '.JPEG', '.jpeg', '.png', '.PNG']
    # get all the file name of images to list
    # r = root, d = directories, f = files
    for r, d, f in os.walk(Input_Folder_Path):
        for file_name in f:
            if any(Each_Image_Extension in file_name for Each_Image_Extension in Image_Extention):
                Files_Name_List.append(file_name)
    # print(Files_Name_List)

    for File_Name_Idx, File_Name in enumerate(Files_Name_List):
        cv2.waitKey(100)
        print("---------------------------------------------------------------------")
        # load image and convert to grayscale
        print('File Name:\t{}\t{}/{}'.format(File_Name, File_Name_Idx + 1, len(Files_Name_List)))
        Original_Image = cv2.imread(Input_Folder_Path + File_Name)
        # display image
        cv2.imshow('Original_Image', Original_Image)

        # create trackbars
        cv2.namedWindow("TrackBars")
        cv2.resizeWindow("TrackBars", 640, 240)
        # trackbar for threshold canny
        cv2.createTrackbar("ThresholdCanny1", "TrackBars", 0, L - 1, empty)
        cv2.createTrackbar("ThresholdCanny2", "TrackBars", 0, L - 1, empty)
        cv2.createTrackbar("inRange1", "TrackBars", 0, L - 1, empty)
        cv2.createTrackbar("inRange2", "TrackBars", 0, L - 1, empty)
        cv2.waitKey(100)

        # press key to next processing
        print("Press 'Enter' key")
        cv2.waitKey(0)
        print("Enter 'n' key for next image")
        print("Enter 'q' key to closing program")
        print("Press 'Enter' key to continuing processing")
        input_key_in_for_loop = input("Press key 1: ")
        if input_key_in_for_loop == "n":
            continue
        elif input_key_in_for_loop == "q":
            break
        elif input_key_in_for_loop == "":
            pass
        
        check = False

        while True:
            ThresholdCanny1 = cv2.getTrackbarPos("ThresholdCanny1", "TrackBars")
            ThresholdCanny2 = cv2.getTrackbarPos("ThresholdCanny2", "TrackBars")
            inRange1 = cv2.getTrackbarPos("inRange1", "TrackBars")
            inRange2 = cv2.getTrackbarPos("inRange2", "TrackBars")

            # convert to grayscale
            GrayScale_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2GRAY)
            # display image
            cv2.imshow('GrayScale_Image', GrayScale_Image)
            
            M = GrayScale_Image.shape[0] #Height of image
            N = GrayScale_Image.shape[1] #Width of image

            # smooth image and reduce noise of background
            Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
            # display image
            cv2.imshow('Blur_Image', Blur_Image)
            
            # inRange
            inRange_Image = cv2.inRange(Blur_Image, inRange1, inRange2)
            # display image
            cv2.imshow('inRange_Image', inRange_Image)

            # using canny algorithm to find out the edge
            Canny_Image = cv2.Canny(inRange_Image, ThresholdCanny1, ThresholdCanny2)
            # display image
            cv2.imshow('Canny_Image', Canny_Image)

            # using dilate operation to try to find connected edge components
            Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
            # display image
            cv2.imshow('Dilate', Dilate_Image)

            if check == False:    
                print("Enter 'c' key to commanding")
                check = True
            if cv2.waitKey(1) & 0xFF == ord('c'):
                print("Enter 'n' key for next image")
                print("Press 'Enter' key to continuing processing")
                input_key_in_while_true = input("Press key 2: ")
                if input_key_in_while_true == "n":
                    cv2.destroyAllWindows()
                    break
                if input_key_in_while_true == "":
                    print("ThresholdCanny1: " + str(ThresholdCanny1))
                    print("ThresholdCanny2: " + str(ThresholdCanny2))
                    print("inRange1: " + str(inRange1))
                    print("inRange2: " + str(inRange2))
                    # using FloodFill to remove noise
                    FloodFill_Image = Dilate_Image
                    Gray_FloodFill = Gray_FloodFill_Ini
                    Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
                    count = 0
                    for x in range(0, M):
                        for y in range(0, N):
                            r = Dilate_Image[x, y]
                            if (r == L - 1):
                                count += 1
                                cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                                Gray_FloodFill += Gray_FloodFill_Increase_Step
                    # display image
                    cv2.imshow('FloodFill', FloodFill_Image)
                    FloodFill_Area_Components = [0] * L
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r > 0):
                                FloodFill_Area_Components[r] += 1
                    FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
                    # sorting 'FloodFill_Area_Components_Enumerate' from max to min
                    FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
                    print("FloodFill Area Components MaxMin: \n" + str(FloodFill_Area_Components_MaxMin))
                    rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
                    RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # remove unnecessary components (noise)
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r == rmax):
                                RemoveNoise_Image[x, y] = L - 1
                            else:
                                RemoveNoise_Image[x, y] = 0
                    # display image
                    cv2.imshow('RemoveNoise', RemoveNoise_Image)

                    # find contours
                    """
                    Because canny edge detector separate different segments,
                    so, using 'findContours' to assemble those edge pixels.
                    NOTE: The function 'findContours' computes contours from 
                            binary images, so it take images created by 'Canny'
                    """
                    _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    # find fruit contour
                    Number_of_Contours = len(Contours)
                    # using bubble sort to sorting contours from max to min
                    Contours_Sorting_MaxMin = Contours.copy()
                    for i in range(Number_of_Contours):
                        for j in range(0, Number_of_Contours - i - 1):
                            if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                                Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
                    print("Contours_Sorting_MaxMin: " + str(len(Contours_Sorting_MaxMin[0])))
                    # draw fruit contour '0'
                    DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
                    # display image
                    cv2.imshow('DrawFruitContour', DrawFruitContour_Image)

                    # estimate fruit area
                    Fruit_Area = 0
                    Fruit_Perimeter = 0
                    Fruit_Area_Contour = 0
                    Fruit_Perimeter = cv2.arcLength(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], True)
                    Fruit_Area_Contour = cv2.contourArea(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # throughout the pixel image to check it belongs inside contour or not
                    for x in range (0, M):
                        for y in range (0, N):
                            # using pointPolygonTest to testing if a point is inside contour.
                            # the return values are simply +1, -1, or 0 
                            # depending on whether the point is inside, out-side, 
                            # or on an edge.
                            if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                                Fruit_Area += 1
                                Area_Image[x, y] = L - 1
                    print("Fruit_Area: " + str(Fruit_Area))
                    print("Fruit_Area_Contour: " + str(Fruit_Area_Contour))
                    print("Fruit_Perimeter: " + str(Fruit_Perimeter))
                    # display image
                    cv2.imshow('Area', Area_Image)

                    # Measure fruit dimensions
                    # drawing rectangle around contour
                    Dimensions_Image = DrawFruitContour_Image
                    (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
                    print("W_Fruit_Contour: " + str(W_Fruit_Contour))
                    print("H_Fruit_Contour: " + str(H_Fruit_Contour))
                    # display image
                    cv2.imshow('Dimensions_Image', Dimensions_Image)
                    
                    print("Press 'Enter' key")
                    cv2.waitKey(0)
                    cv2.destroyAllWindows()
                    break
                
            cv2.waitKey(1)

# Loop, inRange, excel
def inRange_Loop_Excel():
     # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))

    # inRange
    inRange1 = 14
    inRange2 = 250
    print("inRange1: " + str(inRange1))
    print("inRange2: " + str(inRange2))

    # initialise two thresholds, a lower (ThresholdCanny1) 
    # and an upper (ThresholdCanny2) of canny edge detector.
    """
    If a pixel has a gradient larger than the upper threshold,
    then it is accepted as an edge pixel; if a pixel is below
    the lower threshold, it is rejected. 
    If the pixel’s gradient is between the thresholds, 
    then it will be accepted only if it is connected to 
    a pixel that is above the high threshold.
    NOTE: Canny recommended a ratio of high:low threshold
            between 2:1 and 3:1
    """
    ThresholdCanny1 = 0
    ThresholdCanny2 = 3 * ThresholdCanny1
    print("Threshold Canny 1: " + str(ThresholdCanny1))
    print("Threshold Canny 2: " + str(ThresholdCanny2))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))
    cv2.waitKey(100)

    # Input and output image folder path
    Input_Folder_Path = "E:/CVS/Project/Fruits_Classification/dataset/DB_Royal/"
    Files_Name_List = []
    # the list of image file extensions
    Image_Extention = ['.JPG', '.jpg', '.JPEG', '.jpeg', '.png', '.PNG']
    # r = root, d = directories, f = files
    for r, d, f in os.walk(Input_Folder_Path):
        for file_name in f:
            if any(Each_Image_Extension in file_name for Each_Image_Extension in Image_Extention):
                Files_Name_List.append(file_name)
    # print(Files_Name_List)
    Output_Folder_Gray_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/gray/"
    Output_Folder_Blur_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/blur/"
    Output_Folder_inRange_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/inrange/"
    Output_Folder_Canny_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/canny/"
    Output_Folder_Dilate_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/dilate/"
    Output_Folder_FloodFill_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/floodfill/"
    Output_Folder_RemoveNoise_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/remove_noise/"
    Output_Folder_Contour_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/contour/"
    Output_Folder_Area_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/area/"
    Output_Folder_Isolate_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/isolate/"
    Output_Folder_Dimension_Path = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/dimension/"

    # Excel file path
    Output_Folder_Excel_File = "E:/CVS/Project/Fruits_Classification/image_result/inRange/DB_Royal4/"
    
    # create workbook
    WorkBook = xlsxwriter.Workbook(Output_Folder_Excel_File + "DB_Royal4_inRange_14_255.xlsx")
    
    # create worksheet for writing parameters
    Parameters_WorkSheet = WorkBook.add_worksheet("Parameters")
    # write parameters to worksheet
    Parameters_WorkSheet.write('A1', "Blur Kernel Size: ")
    Parameters_WorkSheet.write('A2', "inRange1: ") 
    Parameters_WorkSheet.write('A3', "inRange2: ") 
    Parameters_WorkSheet.write('A4', "Threshold Canny 1: ") 
    Parameters_WorkSheet.write('A5', "Threshold Canny 2: ") 
    Parameters_WorkSheet.write('A6', "Dilate Kernel Size: ") 
    Parameters_WorkSheet.write('A7', "Gray FloodFill: ") 
    Parameters_WorkSheet.write('A8', "Gray FloodFill Increase Num: ") 
    Parameters_WorkSheet.write('A9', "FloodFill Area Components MaxMin Idx 0: ") 
    Parameters_WorkSheet.write('A10', "FloodFill Area Components MaxMin Idx 1: ") 
    Parameters_WorkSheet.write('A11', "Fruit Contours Sorting MaxMin Idx: ") 
    Parameters_Excel_File_List = []
    Parameters_Excel_File_List.append(str(Blur_KernelSize))
    Parameters_Excel_File_List.append(str(inRange1))
    Parameters_Excel_File_List.append(str(inRange2))
    Parameters_Excel_File_List.append(str(ThresholdCanny1))
    Parameters_Excel_File_List.append(str(ThresholdCanny2))
    Parameters_Excel_File_List.append(str(Dilate_KernelSize))
    Parameters_Excel_File_List.append(str(Gray_FloodFill_Ini))
    Parameters_Excel_File_List.append(str(Gray_FloodFill_Increase_Step))
    Parameters_Excel_File_List.append(str(FloodFill_Area_Components_MaxMin_Idx0))
    Parameters_Excel_File_List.append(str(FloodFill_Area_Components_MaxMin_Idx1))
    Parameters_Excel_File_List.append(str(Fruit_Contours_Sorting_MaxMin_Idx))
    # print(Parameters_Excel_File_List)
    Parameters_Row = 0
    Parameters_Column = 1
    for item in Parameters_Excel_File_List:
        # write operation perform 
        Parameters_WorkSheet.write(Parameters_Row, Parameters_Column, item) 
        # incrementing the value of row by one with each iteratons. 
        Parameters_Row += 1 

    # create worksheet for writing measurements
    Measurements_WorkSheet = WorkBook.add_worksheet("Measurements")
    Measurements_WorkSheet.write('B1', "Time")
    Measurements_WorkSheet.write('C1', "Area")
    Measurements_WorkSheet.write('D1', "Contour_Area")
    Measurements_WorkSheet.write('E1', "Perimeter")
    Measurements_WorkSheet.write('F1', "Width")
    Measurements_WorkSheet.write('G1', "Height")
    Measurements_WorkSheet.write('H1', "FloodFill_Area_Components_MaxMin")
    Measurements_WorkSheet.write('I1', "Contours_Sorting_MaxMin")

    for File_Name_Idx, File_Name in enumerate(Files_Name_List):
        # load image
        print('File Name:\t{}\t{}/{}'.format(File_Name, File_Name_Idx + 1, len(Files_Name_List)))
        # the timing of processing
        Start_Time = timer()
        Original_Image = cv2.imread(Input_Folder_Path + File_Name)

        # convert to grayscale
        GrayScale_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2GRAY)
        # save image
        cv2.imwrite(Output_Folder_Gray_Path + File_Name, GrayScale_Image)

        M = GrayScale_Image.shape[0] #Height of image
        N = GrayScale_Image.shape[1] #Width of image

        # smooth image and reduce noise of background
        Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
        # save image
        cv2.imwrite(Output_Folder_Blur_Path + File_Name, Blur_Image)
        
        # inRange
        inRange_Image = cv2.inRange(Blur_Image, inRange1, inRange2)
        # save image
        cv2.imwrite(Output_Folder_inRange_Path + File_Name, inRange_Image)

        # using canny algorithm to find out the edge
        Canny_Image = cv2.Canny(inRange_Image, ThresholdCanny1, ThresholdCanny2)
        # save image
        cv2.imwrite(Output_Folder_Canny_Path + File_Name, Canny_Image)

        # using dilate operation to try to find connected edge components
        Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
        # save image
        cv2.imwrite(Output_Folder_Dilate_Path + File_Name, Dilate_Image)

        # using FloodFill to remove noise
        FloodFill_Image = Dilate_Image
        Gray_FloodFill = Gray_FloodFill_Ini
        Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
        count = 0
        for x in range(0, M):
            for y in range(0, N):
                r = Dilate_Image[x, y]
                if (r == L - 1):
                    count += 1
                    cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                    Gray_FloodFill += Gray_FloodFill_Increase_Step
        # save image
        cv2.imwrite(Output_Folder_FloodFill_Path + File_Name, FloodFill_Image)
        FloodFill_Area_Components = [0] * L
        for x in range(0, M):
            for y in range(0, N):
                r = FloodFill_Image[x, y]
                if (r > 0):
                    FloodFill_Area_Components[r] += 1
        FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
        # sorting 'FloodFill_Area_Components_Enumerate' from max to min
        FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
        rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
        RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        for x in range(0, M):
            for y in range(0, N):
                r = FloodFill_Image[x, y]
                if (r == rmax):
                    RemoveNoise_Image[x, y] = L - 1
                else:
                    RemoveNoise_Image[x, y] = 0
        # save image
        cv2.imwrite(Output_Folder_RemoveNoise_Path + File_Name, RemoveNoise_Image)

        # find contours
        """
        Because canny edge detector separate different segments,
        so, using 'findContours' to assemble those edge pixels.
        NOTE: The function 'findContours' computes contours from 
                binary images, so it take images created by 'Canny'
        """
        _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # find fruit contour
        Number_of_Contours = len(Contours)
        # using bubble sort to sorting contours from max to min
        Contours_Sorting_MaxMin = Contours.copy()
        for i in range(Number_of_Contours):
            for j in range(0, Number_of_Contours - i - 1):
                if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                    Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
        # draw fruit contour '0'
        DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
        # save image
        cv2.imwrite(Output_Folder_Contour_Path + File_Name, DrawFruitContour_Image)

        # estimate fruit area
        Fruit_Area = 0
        Fruit_Perimeter = 0
        Fruit_Area_Contour = 0
        Fruit_Perimeter = cv2.arcLength(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], True)
        Fruit_Area_Contour = cv2.contourArea(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
        Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        Isolate_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
        # throughout the pixel image to check it belongs inside contour or not
        for x in range (0, M):
            for y in range (0, N):
                # using pointPolygonTest to testing if a point is inside contour.
                # the return values are simply +1, -1, or 0 
                # depending on whether the point is inside, out-side, 
                # or on an edge.
                if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                    Fruit_Area += 1
                    Area_Image[x, y] = L - 1
                    Isolate_Image[x, y] = GrayScale_Image[x, y]
        # save image
        cv2.imwrite(Output_Folder_Area_Path + File_Name, Area_Image)
        cv2.imwrite(Output_Folder_Isolate_Path + File_Name, Isolate_Image)

        # Measure fruit dimensions
        # drawing rectangle around contour
        Dimensions_Image = DrawFruitContour_Image
        (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
        cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
        # save image
        cv2.imwrite(Output_Folder_Dimension_Path + File_Name, Dimensions_Image)

        Stop_Time = timer() - Start_Time

        # write measurements to worksheet
        Measurements_WorkSheet.write(File_Name_Idx + 1, 0, File_Name)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 1, Stop_Time)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 2, Fruit_Area)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 3, Fruit_Area_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 4, Fruit_Perimeter)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 5, W_Fruit_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 6, H_Fruit_Contour)
        Measurements_WorkSheet.write(File_Name_Idx + 1, 7, str(FloodFill_Area_Components_MaxMin))
        Measurements_WorkSheet.write(File_Name_Idx + 1, 8, str(len(Contours_Sorting_MaxMin[0]))) 

        cv2.waitKey(1)
    
    # close excel file
    WorkBook.close()
########################################################

# Loop, threshold, Trackbars
def Threshold_Loop_Trackbars():
    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))
    cv2.waitKey(100)

    # Input image folder path
    Input_Folder_Path = "E:/CVS/Project/Fruits_Classification/dataset/Mango/"
    Files_Name_List = []
    # the list of image file extensions
    Image_Extention = ['.JPG', '.jpg', '.JPEG', '.jpeg', '.png', '.PNG']
    # get all the file name of images to list
    # r = root, d = directories, f = files
    for r, d, f in os.walk(Input_Folder_Path):
        for file_name in f:
            if any(Each_Image_Extension in file_name for Each_Image_Extension in Image_Extention):
                Files_Name_List.append(file_name)
    # print(Files_Name_List)

    for File_Name_Idx, File_Name in enumerate(Files_Name_List):
        cv2.waitKey(100)
        print("---------------------------------------------------------------------")
        # load image and convert to grayscale
        print('File Name:\t{}\t{}/{}'.format(File_Name, File_Name_Idx + 1, len(Files_Name_List)))
        Original_Image = cv2.imread(Input_Folder_Path + File_Name)
        # display image
        cv2.imshow('Original_Image', Original_Image)

        # create trackbars
        cv2.namedWindow("TrackBars")
        cv2.resizeWindow("TrackBars", 640, 240)
        # trackbar for threshold canny
        cv2.createTrackbar("ThresholdCanny1", "TrackBars", 0, L - 1, empty)
        cv2.createTrackbar("ThresholdCanny2", "TrackBars", 0, L - 1, empty)
        cv2.createTrackbar("Threshold1", "TrackBars", 0, L - 1, empty)
        cv2.waitKey(100)

        # press key to next processing
        print("Press 'Enter' key")
        cv2.waitKey(0)
        print("Enter 'n' key for next image")
        print("Enter 'q' key to closing program")
        print("Press 'Enter' key to continuing processing")
        input_key_in_for_loop = input("Press key 1: ")
        if input_key_in_for_loop == "n":
            continue
        elif input_key_in_for_loop == "q":
            break
        elif input_key_in_for_loop == "":
            pass
        
        check = False

        while True:
            ThresholdCanny1 = cv2.getTrackbarPos("ThresholdCanny1", "TrackBars")
            ThresholdCanny2 = cv2.getTrackbarPos("ThresholdCanny2", "TrackBars")
            Threshold1 = cv2.getTrackbarPos("Threshold1", "TrackBars")

            # convert to grayscale
            GrayScale_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2GRAY)
            # display image
            cv2.imshow('GrayScale_Image', GrayScale_Image)
            
            M = GrayScale_Image.shape[0] #Height of image
            N = GrayScale_Image.shape[1] #Width of image

            # smooth image and reduce noise of background
            Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
            # display image
            cv2.imshow('Blur_Image', Blur_Image)
            
            # Threshold
            ret, Threshold_Image = cv2.threshold(Blur_Image, Threshold1, 255, cv2.THRESH_BINARY)
            # display image
            cv2.imshow('Threshold_Image', Threshold_Image)

            # using canny algorithm to find out the edge
            Canny_Image = cv2.Canny(Threshold_Image, ThresholdCanny1, ThresholdCanny2)
            # display image
            cv2.imshow('Canny_Image', Canny_Image)

            # using dilate operation to try to find connected edge components
            Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
            # display image
            cv2.imshow('Dilate', Dilate_Image)

            if check == False:    
                print("Enter 'c' key to commanding")
                check = True
            if cv2.waitKey(1) & 0xFF == ord('c'):
                print("Enter 'n' key for next image")
                print("Press 'Enter' key to continuing processing")
                input_key_in_while_true = input("Press key 2: ")
                if input_key_in_while_true == "n":
                    cv2.destroyAllWindows()
                    break
                if input_key_in_while_true == "":
                    print("ThresholdCanny1: " + str(ThresholdCanny1))
                    print("ThresholdCanny2: " + str(ThresholdCanny2))
                    print("Threshold1: " + str(Threshold1))
                    # using FloodFill to remove noise
                    FloodFill_Image = Dilate_Image
                    Gray_FloodFill = Gray_FloodFill_Ini
                    Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
                    count = 0
                    for x in range(0, M):
                        for y in range(0, N):
                            r = Dilate_Image[x, y]
                            if (r == L - 1):
                                count += 1
                                cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                                Gray_FloodFill += Gray_FloodFill_Increase_Step
                    # display image
                    cv2.imshow('FloodFill', FloodFill_Image)
                    FloodFill_Area_Components = [0] * L
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r > 0):
                                FloodFill_Area_Components[r] += 1
                    FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
                    # sorting 'FloodFill_Area_Components_Enumerate' from max to min
                    FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
                    print("FloodFill Area Components MaxMin: \n" + str(FloodFill_Area_Components_MaxMin))
                    rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
                    RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # remove unnecessary components (noise)
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r == rmax):
                                RemoveNoise_Image[x, y] = L - 1
                            else:
                                RemoveNoise_Image[x, y] = 0
                    # display image
                    cv2.imshow('RemoveNoise', RemoveNoise_Image)

                    # find contours
                    """
                    Because canny edge detector separate different segments,
                    so, using 'findContours' to assemble those edge pixels.
                    NOTE: The function 'findContours' computes contours from 
                            binary images, so it take images created by 'Canny'
                    """
                    _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    # find fruit contour
                    Number_of_Contours = len(Contours)
                    # using bubble sort to sorting contours from max to min
                    Contours_Sorting_MaxMin = Contours.copy()
                    for i in range(Number_of_Contours):
                        for j in range(0, Number_of_Contours - i - 1):
                            if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                                Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
                    print("Contours_Sorting_MaxMin: " + str(len(Contours_Sorting_MaxMin[0])))
                    # draw fruit contour '0'
                    DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
                    # display image
                    cv2.imshow('DrawFruitContour', DrawFruitContour_Image)

                    # estimate fruit area
                    Fruit_Area = 0
                    Fruit_Perimeter = 0
                    Fruit_Area_Contour = 0
                    Fruit_Perimeter = cv2.arcLength(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], True)
                    Fruit_Area_Contour = cv2.contourArea(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # throughout the pixel image to check it belongs inside contour or not
                    for x in range (0, M):
                        for y in range (0, N):
                            # using pointPolygonTest to testing if a point is inside contour.
                            # the return values are simply +1, -1, or 0 
                            # depending on whether the point is inside, out-side, 
                            # or on an edge.
                            if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                                Fruit_Area += 1
                                Area_Image[x, y] = L - 1
                    print("Fruit_Area: " + str(Fruit_Area))
                    print("Fruit_Area_Contour: " + str(Fruit_Area_Contour))
                    print("Fruit_Perimeter: " + str(Fruit_Perimeter))
                    # display image
                    cv2.imshow('Area', Area_Image)

                    # Measure fruit dimensions
                    # drawing rectangle around contour
                    Dimensions_Image = DrawFruitContour_Image
                    (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
                    print("W_Fruit_Contour: " + str(W_Fruit_Contour))
                    print("H_Fruit_Contour: " + str(H_Fruit_Contour))
                    # display image
                    cv2.imshow('Dimensions_Image', Dimensions_Image)
                    
                    print("Press 'Enter' key")
                    cv2.waitKey(0)
                    cv2.destroyAllWindows()
                    break
                
            cv2.waitKey(1)
########################################################

# Trackbars, Manual, excel
def Trackbars_Manual_RGB_inRange_HSV_Threshold():
    # the number of possible intensity levels in the image 
    # (256 for an 8-bit image).
    L = 256
    # initialise the kernel size of 'blur' function.
    """
    it means each pixel in the output is the simple mean
    of all of the pixels in a window (the kernel),
    around the corresponding pixel in the input.
    """
    Blur_KernelSize = (5, 5)
    print("Blur kernel size: " + str(Blur_KernelSize))
    # initialise the kernel of dilation function
    """
    In dilation function, any given pixel is replaced with
    the local maximum of all of the pixel values covered by
    the kernel.
    """
    Dilate_KernelSize = np.ones((3, 3), np.uint8)
    print("Dilate Kernel Size: " + str(Dilate_KernelSize))
    # flood fill
    Gray_FloodFill_Ini = 10
    print("Gray FloodFill: " + str(Gray_FloodFill_Ini))
    Gray_FloodFill_Increase_Step = 1
    print("Gray FloodFill Increase Num: " + str(Gray_FloodFill_Increase_Step))
    FloodFill_Area_Components_MaxMin_Idx0 = 0
    print("FloodFill Area Components MaxMin Idx 0: " + str(FloodFill_Area_Components_MaxMin_Idx0))
    FloodFill_Area_Components_MaxMin_Idx1 = 0
    print("FloodFill Area Components MaxMin Idx 1: " + str(FloodFill_Area_Components_MaxMin_Idx1))
    # contour
    Fruit_Contours_Sorting_MaxMin_Idx = 0
    print("Fruit Contours Sorting MaxMin Idx: " + str(Fruit_Contours_Sorting_MaxMin_Idx))
    cv2.waitKey(100)

    # Input image folder path
    Input_Folder_Path = "E:/CVS/Project/Fruits_Classification/dataset/Mango/"
    Files_Name_List = []
    # the list of image file extensions
    Image_Extention = ['.JPG', '.jpg', '.JPEG', '.jpeg', '.png', '.PNG']
    # get all the file name of images to list
    # r = root, d = directories, f = files
    for r, d, f in os.walk(Input_Folder_Path):
        for file_name in f:
            if any(Each_Image_Extension in file_name for Each_Image_Extension in Image_Extention):
                Files_Name_List.append(file_name)
    # print(Files_Name_List)

    # output result path
    Fruit_Name_OutputPath = "Mango/"
    BeginOutputPath = "../result/"
    Fruit_Name_OutputDirectory = os.path.join(BeginOutputPath, Fruit_Name_OutputPath)
    os.mkdir(Fruit_Name_OutputDirectory)    # create the result directory of fruit name
    # Hsv method output directory
    HSV_method_OutputPath = "HSV_method/"
    HSV_method_OutputDirectory = os.path.join(Fruit_Name_OutputDirectory, HSV_method_OutputPath)
    os.mkdir(HSV_method_OutputDirectory)    # create the result directory of the HSV processing 
    HSV_method_hsv_Folder_OutputPath = "hsv/"
    HSV_method_hsv_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_hsv_Folder_OutputPath)
    os.mkdir(HSV_method_hsv_Folder_OutputDirectory) # create 'hsv' folder directory
    HSV_method_gray_Folder_OutputPath = "gray/"
    HSV_method_gray_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_gray_Folder_OutputPath)
    os.mkdir(HSV_method_gray_Folder_OutputDirectory)    # create 'gray' folder directory
    HSV_method_blur_Folder_OutputPath = "blur/"
    HSV_method_blur_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_blur_Folder_OutputPath)
    os.mkdir(HSV_method_blur_Folder_OutputDirectory)    # create 'blur' folder directory
    HSV_method_canny_Folder_OutputPath = "canny/"
    HSV_method_canny_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_canny_Folder_OutputPath)
    os.mkdir(HSV_method_canny_Folder_OutputDirectory)   # create 'canny' folder directory
    HSV_method_dilate_Folder_OutputPath = "dilate/"
    HSV_method_dilate_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_dilate_Folder_OutputPath)
    os.mkdir(HSV_method_dilate_Folder_OutputDirectory)  # create 'dilate' folder directory
    HSV_method_floodfill_Folder_OutputPath = "floodfill/"
    HSV_method_floodfill_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_floodfill_Folder_OutputPath)
    os.mkdir(HSV_method_floodfill_Folder_OutputDirectory)   # create 'floodfill' folder directory
    HSV_method_remove_noise_Folder_OutputPath = "remove_noise/"
    HSV_method_remove_noise_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_remove_noise_Folder_OutputPath)
    os.mkdir(HSV_method_remove_noise_Folder_OutputDirectory)    # create 'remove_noise' folder directory
    HSV_method_contour_Folder_OutputPath = "contour/"
    HSV_method_contour_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_contour_Folder_OutputPath)
    os.mkdir(HSV_method_contour_Folder_OutputDirectory)
    HSV_method_area_Folder_OutputPath = "area/"
    HSV_method_area_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_area_Folder_OutputPath)
    os.mkdir(HSV_method_area_Folder_OutputDirectory)
    HSV_method_dimension_Folder_OutputPath = "dimension/"
    HSV_method_dimension_Folder_OutputDirectory = os.path.join(HSV_method_OutputDirectory, HSV_method_dimension_Folder_OutputPath)
    os.mkdir(HSV_method_dimension_Folder_OutputDirectory)

    # Excel file path
    ExcelFile_OutputPath = HSV_method_OutputDirectory
    
    # create workbook
    WorkBook = xlsxwriter.Workbook(ExcelFile_OutputPath + "HSV_Mango.xlsx")
    
    # create worksheet for writing parameters
    Parameters_WorkSheet = WorkBook.add_worksheet("Parameters")
    # write parameters to worksheet
    Parameters_WorkSheet.write('A1', "Blur Kernel Size: ") 
    Parameters_WorkSheet.write('A2', "Dilate Kernel Size: ") 
    Parameters_WorkSheet.write('A3', "Gray FloodFill: ") 
    Parameters_WorkSheet.write('A4', "Gray FloodFill Increase Num: ") 
    Parameters_WorkSheet.write('A5', "FloodFill Area Components MaxMin Idx 0: ") 
    Parameters_WorkSheet.write('A6', "FloodFill Area Components MaxMin Idx 1: ") 
    Parameters_WorkSheet.write('A7', "Fruit Contours Sorting MaxMin Idx: ") 
    Parameters_Excel_File_List = []
    Parameters_Excel_File_List.append(str(Blur_KernelSize))
    Parameters_Excel_File_List.append(str(Dilate_KernelSize))
    Parameters_Excel_File_List.append(str(Gray_FloodFill_Ini))
    Parameters_Excel_File_List.append(str(Gray_FloodFill_Increase_Step))
    Parameters_Excel_File_List.append(str(FloodFill_Area_Components_MaxMin_Idx0))
    Parameters_Excel_File_List.append(str(FloodFill_Area_Components_MaxMin_Idx1))
    Parameters_Excel_File_List.append(str(Fruit_Contours_Sorting_MaxMin_Idx))
    # print(Parameters_Excel_File_List)
    Parameters_Row = 0
    Parameters_Column = 1
    for item in Parameters_Excel_File_List:
        # write operation perform 
        Parameters_WorkSheet.write(Parameters_Row, Parameters_Column, item) 
        # incrementing the value of row by one with each iteratons. 
        Parameters_Row += 1 

    Row_Measurements_WorkSheet = 1

    # create worksheet for writing measurements
    Measurements_WorkSheet = WorkBook.add_worksheet("Measurements")
    Measurements_WorkSheet.write('B1', "Area")
    Measurements_WorkSheet.write('C1', "Contour_Area")
    Measurements_WorkSheet.write('D1', "Perimeter")
    Measurements_WorkSheet.write('E1', "Width")
    Measurements_WorkSheet.write('F1', "Height")
    Measurements_WorkSheet.write('G1', "FloodFill_Area_Components_MaxMin")
    Measurements_WorkSheet.write('H1', "Contours_Sorting_MaxMin")
    Measurements_WorkSheet.write('I1', "ThresholdCanny1")
    Measurements_WorkSheet.write('J1', "ThresholdCanny2")

    for File_Name_Idx, File_Name in enumerate(Files_Name_List):
        cv2.waitKey(100)
        print("---------------------------------------------------------------------")
        # load image and convert to grayscale
        print('File Name:\t{}\t{}/{}'.format(File_Name, File_Name_Idx + 1, len(Files_Name_List)))
        Original_Image = cv2.imread(Input_Folder_Path + File_Name)
        # display image
        cv2.imshow('Original_Image', Original_Image)

        # create trackbars
        cv2.namedWindow("TrackBars")
        cv2.resizeWindow("TrackBars", 640, 240)
        # trackbar for threshold canny
        cv2.createTrackbar("ThresholdCanny1", "TrackBars", 0, L - 1, empty)
        cv2.createTrackbar("ThresholdCanny2", "TrackBars", 0, L - 1, empty)

        cv2.waitKey(100)

        # press key to next processing
        print("Enter 'n' key for next image")
        print("Enter 'q' key to closing program")
        print("Not enter any key to continuing processing")
        input_key_in_for_loop = input("Press key 1: ")
        if input_key_in_for_loop == "n":
            continue
        elif input_key_in_for_loop == "q":
            break
        elif input_key_in_for_loop == "":
            pass
        
        check = False
        while True:
            # convert BGR to HSV
            HSV_Image = cv2.cvtColor(Original_Image, cv2.COLOR_BGR2HSV)
            # display image
            cv2.imshow('HSV_Image', HSV_Image)

            # convert to grayscale
            GrayScale_Image = cv2.cvtColor(HSV_Image, cv2.COLOR_BGR2GRAY)
            # display image
            cv2.imshow('GrayScale_Image', GrayScale_Image)
            
            M = GrayScale_Image.shape[0] #Height of image
            N = GrayScale_Image.shape[1] #Width of image

            # smooth image and reduce noise of background
            Blur_Image = cv2.blur(GrayScale_Image, Blur_KernelSize)
            # display image
            cv2.imshow('Blur_Image', Blur_Image)
            
            # using canny algorithm to find out the edge
            ThresholdCanny1 = cv2.getTrackbarPos("ThresholdCanny1", "TrackBars")
            ThresholdCanny2 = cv2.getTrackbarPos("ThresholdCanny2", "TrackBars")
            Canny_Image = cv2.Canny(Blur_Image, ThresholdCanny1, ThresholdCanny2)
            # display image
            cv2.imshow('Canny_Image', Canny_Image)

            # using dilate operation to try to find connected edge components
            Dilate_Image = cv2.dilate(Canny_Image, Dilate_KernelSize, iterations=1)
            # display image
            cv2.imshow('Dilate', Dilate_Image)

            if check == False:    
                print("Enter 'c' key to commanding")
                check = True
            if cv2.waitKey(1) & 0xFF == ord('c'):
                print("Enter 'n' key for next image")
                print("Not enter any key to continuing processing")
                input_key_in_while_true = input("Press key 2: ")
                if input_key_in_while_true == "n":
                    cv2.destroyAllWindows()
                    break
                if input_key_in_while_true == "":
                    # using FloodFill to remove noise
                    FloodFill_Image = Dilate_Image
                    Gray_FloodFill = Gray_FloodFill_Ini
                    Mask_FloodFill = np.zeros((M+2, N+2), np.uint8)
                    count = 0
                    for x in range(0, M):
                        for y in range(0, N):
                            r = Dilate_Image[x, y]
                            if (r == L - 1):
                                count += 1
                                cv2.floodFill(FloodFill_Image, Mask_FloodFill, (y, x), Gray_FloodFill)
                                Gray_FloodFill += Gray_FloodFill_Increase_Step
                    # display image
                    cv2.imshow('FloodFill', FloodFill_Image)
                    FloodFill_Area_Components = [0] * L
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r > 0):
                                FloodFill_Area_Components[r] += 1
                    FloodFill_Area_Components_Enumerate = list(enumerate(FloodFill_Area_Components))
                    # sorting 'FloodFill_Area_Components_Enumerate' from max to min
                    FloodFill_Area_Components_MaxMin = sorted(FloodFill_Area_Components_Enumerate, key=lambda x:x [1], reverse = True)
                    rmax = FloodFill_Area_Components_MaxMin[FloodFill_Area_Components_MaxMin_Idx0][FloodFill_Area_Components_MaxMin_Idx1]
                    RemoveNoise_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # remove unnecessary components (noise)
                    for x in range(0, M):
                        for y in range(0, N):
                            r = FloodFill_Image[x, y]
                            if (r == rmax):
                                RemoveNoise_Image[x, y] = L - 1
                            else:
                                RemoveNoise_Image[x, y] = 0
                    # display image
                    cv2.imshow('RemoveNoise', RemoveNoise_Image)

                    # find contours
                    """
                    Because canny edge detector separate different segments,
                    so, using 'findContours' to assemble those edge pixels.
                    NOTE: The function 'findContours' computes contours from 
                            binary images, so it take images created by 'Canny'
                    """
                    _, Contours, Hierarchy = cv2.findContours(RemoveNoise_Image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    # find fruit contour
                    Number_of_Contours = len(Contours)
                    # using bubble sort to sorting contours from max to min
                    Contours_Sorting_MaxMin = Contours.copy()
                    for i in range(Number_of_Contours):
                        for j in range(0, Number_of_Contours - i - 1):
                            if (len(Contours_Sorting_MaxMin[j]) < len(Contours_Sorting_MaxMin[j + 1])):
                                Contours_Sorting_MaxMin[j], Contours_Sorting_MaxMin[j + 1] = Contours_Sorting_MaxMin[j + 1], Contours_Sorting_MaxMin[j]
                    # draw fruit contour '0'
                    DrawFruitContour_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    cv2.drawContours(DrawFruitContour_Image, Contours_Sorting_MaxMin, Fruit_Contours_Sorting_MaxMin_Idx, (255, 0, 0), 1)
                    # display image
                    cv2.imshow('DrawFruitContour', DrawFruitContour_Image)

                    # estimate fruit area
                    Fruit_Area = 0
                    Fruit_Perimeter = 0
                    Fruit_Area_Contour = 0
                    Fruit_Perimeter = cv2.arcLength(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], True)
                    Fruit_Area_Contour = cv2.contourArea(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    Area_Image = np.zeros((GrayScale_Image.shape[0], GrayScale_Image.shape[1], 1), np.uint8)
                    # throughout the pixel image to check it belongs inside contour or not
                    for x in range (0, M):
                        for y in range (0, N):
                            # using pointPolygonTest to testing if a point is inside contour.
                            # the return values are simply +1, -1, or 0 
                            # depending on whether the point is inside, out-side, 
                            # or on an edge.
                            if (cv2.pointPolygonTest(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx], (y, x), False) > 0):
                                Fruit_Area += 1
                                Area_Image[x, y] = L - 1
                    # display image
                    cv2.imshow('Area', Area_Image)

                    # Measure fruit dimensions
                    # drawing rectangle around contour
                    Dimensions_Image = DrawFruitContour_Image
                    (X_Fruit_Contour, Y_Fruit_Contour, W_Fruit_Contour, H_Fruit_Contour) = cv2.boundingRect(Contours_Sorting_MaxMin[Fruit_Contours_Sorting_MaxMin_Idx])
                    cv2.rectangle(Dimensions_Image, (X_Fruit_Contour, Y_Fruit_Contour), (X_Fruit_Contour + W_Fruit_Contour, Y_Fruit_Contour + H_Fruit_Contour), (255, 0, 0), 1)
                    # display image
                    cv2.imshow('Dimensions_Image', Dimensions_Image)
                    
                    print("Press 'Enter' key")
                    cv2.waitKey(0)
                    print("Enter 's' key to saving result, auto next image")
                    print("Enter 'n' key for next image not save result")
                    input_key_in_while_true = input("Press key 3: ")
                    if input_key_in_while_true == "s":
                        # save image
                        cv2.imwrite(HSV_method_hsv_Folder_OutputDirectory + File_Name, HSV_Image)
                        cv2.imwrite(HSV_method_gray_Folder_OutputDirectory + File_Name, GrayScale_Image)
                        cv2.imwrite(HSV_method_blur_Folder_OutputDirectory + File_Name, Blur_Image)
                        cv2.imwrite(HSV_method_canny_Folder_OutputDirectory + File_Name, Canny_Image)
                        cv2.imwrite(HSV_method_dilate_Folder_OutputDirectory + File_Name, Dilate_Image)
                        cv2.imwrite(HSV_method_floodfill_Folder_OutputDirectory + File_Name, FloodFill_Image)
                        cv2.imwrite(HSV_method_remove_noise_Folder_OutputDirectory + File_Name, RemoveNoise_Image)
                        cv2.imwrite(HSV_method_contour_Folder_OutputDirectory + File_Name, DrawFruitContour_Image)
                        cv2.imwrite(HSV_method_area_Folder_OutputDirectory + File_Name, Area_Image)
                        cv2.imwrite(HSV_method_dimension_Folder_OutputDirectory + File_Name, Dimensions_Image)
                        
                        # write measurements to worksheet
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 0, File_Name)
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 1, Fruit_Area)
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 2, Fruit_Area_Contour)
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 3, Fruit_Perimeter)
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 4, W_Fruit_Contour)
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 5, H_Fruit_Contour)
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 6, str(FloodFill_Area_Components_MaxMin))
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 7, str(len(Contours_Sorting_MaxMin[0]))) 
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 8, ThresholdCanny1) 
                        Measurements_WorkSheet.write(Row_Measurements_WorkSheet, 9, ThresholdCanny2) 
                        
                        Row_Measurements_WorkSheet += 1

                        cv2.destroyAllWindows()
                        break
                    if input_key_in_while_true == "n":
                        cv2.destroyAllWindows()
                        break
            cv2.waitKey(1)
    # close excel file
    WorkBook.close()

def empty(a):
    pass

if __name__ == "__main__":
    Fruits_Feature_Extraction_Trackbars_Threshold()
    print("\n\n-------------------------DONE-------------------------\n\n")